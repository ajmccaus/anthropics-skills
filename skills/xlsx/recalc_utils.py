#!/usr/bin/env python3
"""
Efficiency utilities for xlsx recalc operations.
Optimized error detection, cell iteration, and workbook analysis.
"""

import re
from typing import Optional, List, Dict, Any, Iterator, Generator
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


# Compiled regex for Excel error detection - O(1) pattern matching
# Matches: #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A
EXCEL_ERROR_PATTERN = re.compile(r'#(?:VALUE!|DIV/0!|REF!|NAME\?|NULL!|NUM!|N/A)')

# Minimum length of an Excel error string (shortest is #N/A = 4 chars)
MIN_ERROR_LENGTH = 4

# File size threshold for using read_only mode (in MB)
READ_ONLY_THRESHOLD_MB = 5


def find_errors_in_value(value: str) -> List[str]:
    """
    Find all Excel errors in a cell value.

    Uses compiled regex for efficient matching.

    Args:
        value: The cell value to check.

    Returns:
        List of error strings found (e.g., ['#VALUE!', '#REF!']).
    """
    if not isinstance(value, str) or len(value) < MIN_ERROR_LENGTH:
        return []

    return EXCEL_ERROR_PATTERN.findall(value)


def scan_cell_for_errors(value: Any) -> Optional[str]:
    """
    Scan a single cell value for Excel errors.

    Optimized to skip non-string values and short strings quickly.

    Args:
        value: The cell value to scan.

    Returns:
        The first error found, or None if no errors.
    """
    # Skip None values
    if value is None:
        return None

    # Skip numeric values (can't contain errors)
    if isinstance(value, (int, float)):
        return None

    # Skip non-strings
    if not isinstance(value, str):
        return None

    # Skip strings too short to contain an error
    if len(value) < MIN_ERROR_LENGTH:
        return None

    # Use compiled regex for efficient matching
    match = EXCEL_ERROR_PATTERN.search(value)
    return match.group() if match else None


def get_worksheet_bounds(ws: Worksheet) -> Optional[Dict[str, int]]:
    """
    Get the actual data bounds of a worksheet.

    Returns None for empty worksheets.

    Args:
        ws: The worksheet to analyze.

    Returns:
        Dict with min_row, max_row, min_col, max_col, or None if empty.
    """
    if ws.min_row is None or ws.max_row is None:
        return None

    return {
        'min_row': ws.min_row,
        'max_row': ws.max_row,
        'min_col': ws.min_col or 1,
        'max_col': ws.max_col or 1,
    }


def iterate_data_cells(ws: Worksheet) -> Generator:
    """
    Iterate over cells in a worksheet using bounds.

    Only iterates over the actual data range, not empty cells.

    Args:
        ws: The worksheet to iterate.

    Yields:
        Cell objects from the worksheet.
    """
    bounds = get_worksheet_bounds(ws)
    if bounds is None:
        return

    for row in ws.iter_rows(
        min_row=bounds['min_row'],
        max_row=bounds['max_row'],
        min_col=bounds['min_col'],
        max_col=bounds['max_col']
    ):
        for cell in row:
            yield cell


def is_formula(value: Any) -> bool:
    """
    Check if a cell value is a formula.

    Uses efficient startswith check (O(1)).

    Args:
        value: The cell value to check.

    Returns:
        True if the value is a formula string.
    """
    if value is None:
        return False

    if not isinstance(value, str):
        return False

    if len(value) == 0:
        return False

    # O(1) check - only looks at first character
    return value[0] == '='


def should_use_read_only_mode(file_size_mb: float) -> bool:
    """
    Determine if read_only mode should be used based on file size.

    Args:
        file_size_mb: File size in megabytes.

    Returns:
        True if read_only mode is recommended.
    """
    return file_size_mb >= READ_ONLY_THRESHOLD_MB


def analyze_workbook_single_pass(wb: Workbook) -> Dict[str, Any]:
    """
    Analyze a workbook in a single pass for both formulas and errors.

    Avoids loading the workbook twice with different data_only settings.

    Args:
        wb: The workbook to analyze (loaded with data_only=False).

    Returns:
        Dict with 'formula_count', 'errors', and 'error_count' keys.
    """
    result = {
        'formula_count': 0,
        'errors': {},  # error_type -> list of locations
        'error_count': 0,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for cell in iterate_data_cells(ws):
            value = cell.value

            # Check for formula
            if is_formula(value):
                result['formula_count'] += 1

            # Check for error
            error = scan_cell_for_errors(value)
            if error:
                if error not in result['errors']:
                    result['errors'][error] = []
                result['errors'][error].append(f"{sheet_name}!{cell.coordinate}")
                result['error_count'] += 1

    return result


def scan_worksheet_for_errors(ws: Worksheet, sheet_name: str) -> Dict[str, List[str]]:
    """
    Scan a worksheet for Excel errors efficiently.

    Uses bounded iteration and compiled regex.

    Args:
        ws: The worksheet to scan.
        sheet_name: Name of the worksheet (for error location reporting).

    Returns:
        Dict mapping error types to lists of cell locations.
    """
    errors = {}

    for cell in iterate_data_cells(ws):
        error = scan_cell_for_errors(cell.value)
        if error:
            if error not in errors:
                errors[error] = []
            errors[error].append(f"{sheet_name}!{cell.coordinate}")

    return errors


def count_formulas_in_worksheet(ws: Worksheet) -> int:
    """
    Count formulas in a worksheet efficiently.

    Args:
        ws: The worksheet to analyze.

    Returns:
        Number of formula cells.
    """
    count = 0

    for cell in iterate_data_cells(ws):
        if is_formula(cell.value):
            count += 1

    return count
