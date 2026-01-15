#!/usr/bin/env python3
"""
Excel Formula Recalculation Script

Cross-platform support: Windows, macOS, Linux
- Windows: Uses Excel COM if available (preferred), falls back to LibreOffice
- macOS/Linux: Uses LibreOffice
"""

import json
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

# Import cross-platform utilities
from platform_utils import (
    get_libreoffice_macro_dir,
    get_soffice_command,
    run_with_timeout,
    is_excel_com_available,
    recalc_with_excel_com,
)
from recalc_utils import (
    iterate_data_cells,
    scan_cell_for_errors,
    is_formula,
    get_worksheet_bounds,
)


def setup_libreoffice_macro():
    """Setup LibreOffice macro for recalculation if not already configured."""
    try:
        macro_dir = get_libreoffice_macro_dir()
    except EnvironmentError as e:
        print(f"Warning: {e}", file=sys.stderr)
        return False

    macro_file = os.path.join(macro_dir, 'Module1.xba')

    if os.path.exists(macro_file):
        with open(macro_file, 'r') as f:
            if 'RecalculateAndSave' in f.read():
                return True

    # Find soffice command
    soffice = get_soffice_command()
    if not soffice:
        print("Warning: LibreOffice not found. Please install LibreOffice.", file=sys.stderr)
        return False

    if not os.path.exists(macro_dir):
        try:
            run_with_timeout([soffice, '--headless', '--terminate_after_init'], timeout=10)
        except subprocess.TimeoutExpired:
            pass  # Timeout is acceptable for initialization
        os.makedirs(macro_dir, exist_ok=True)

    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''

    try:
        with open(macro_file, 'w') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def _recalc_with_libreoffice(abs_path: str, timeout: int) -> dict:
    """Recalculate using LibreOffice (fallback method)."""
    if not setup_libreoffice_macro():
        return {'error': 'LibreOffice not configured. Install LibreOffice or Microsoft Excel.'}

    soffice = get_soffice_command()
    if not soffice:
        return {'error': 'LibreOffice not found. Install LibreOffice or Microsoft Excel.'}

    cmd = [
        soffice, '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]

    try:
        result = run_with_timeout(cmd, timeout=timeout)
    except subprocess.TimeoutExpired:
        return {'error': f'Recalculation timed out after {timeout} seconds'}

    if result.returncode != 0:
        error_msg = result.stderr or 'Unknown error during recalculation'
        if 'Module1' in error_msg or 'RecalculateAndSave' not in error_msg:
            return {'error': 'LibreOffice macro not configured properly'}
        return {'error': error_msg}

    return {'success': True}


def recalc(filename, timeout=30):
    """
    Recalculate formulas in Excel file and report any errors.

    On Windows: Uses Excel COM if available (faster, more accurate),
                falls back to LibreOffice if Excel not installed.
    On macOS/Linux: Uses LibreOffice.

    Args:
        filename: Path to Excel file
        timeout: Maximum time to wait for recalculation (seconds)

    Returns:
        dict with error locations and counts
    """
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}

    abs_path = str(Path(filename).absolute())

    # Try Excel COM first on Windows (preferred - faster and more accurate)
    recalc_result = None
    if is_excel_com_available():
        recalc_result = recalc_with_excel_com(abs_path)
        if not recalc_result.get('success'):
            # Excel COM failed, will try LibreOffice
            print(f"Excel COM failed: {recalc_result.get('error')}. Trying LibreOffice...",
                  file=sys.stderr)
            recalc_result = None

    # Fall back to LibreOffice if Excel COM not available or failed
    if recalc_result is None:
        recalc_result = _recalc_with_libreoffice(abs_path, timeout)

    if not recalc_result.get('success') and 'error' in recalc_result:
        return recalc_result

    # Check for Excel errors in calculated values
    try:
        # Load with data_only=True to get calculated values (for error detection)
        wb_values = load_workbook(filename, data_only=True)

        error_details = {}
        total_errors = 0

        for sheet_name in wb_values.sheetnames:
            ws = wb_values[sheet_name]

            # Skip empty sheets efficiently
            if get_worksheet_bounds(ws) is None:
                continue

            # Use optimized bounded iteration
            for cell in iterate_data_cells(ws):
                error = scan_cell_for_errors(cell.value)
                if error:
                    if error not in error_details:
                        error_details[error] = []
                    error_details[error].append(f"{sheet_name}!{cell.coordinate}")
                    total_errors += 1

        wb_values.close()

        # Count formulas (separate load with data_only=False)
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0

        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            if get_worksheet_bounds(ws) is None:
                continue
            for cell in iterate_data_cells(ws):
                if is_formula(cell.value):
                    formula_count += 1

        wb_formulas.close()

        # Build result summary
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'total_formulas': formula_count,
            'error_summary': {}
        }

        for err_type, locations in error_details.items():
            result['error_summary'][err_type] = {
                'count': len(locations),
                'locations': locations[:20]
            }

        return result

    except Exception as e:
        return {'error': str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file.")
        print("  - Windows: Uses Excel (if installed), falls back to LibreOffice")
        print("  - macOS/Linux: Uses LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        sys.exit(1)
    
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    
    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()